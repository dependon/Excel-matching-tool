from flask import Flask, render_template, request, redirect, url_for, send_file, session
from flask_socketio import SocketIO, emit, join_room
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import uuid
import shutil
import time
from threading import Thread
import signal

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret!'
socketio = SocketIO(app)

# 创建一个临时目录来存储上传的文件
TEMP_DIR = 'uploads'
os.makedirs(TEMP_DIR, exist_ok=True)

# 用于记录每个会话的最后访问时间
session_last_access = {}

# 用于控制后台线程的标志
running = True


def cleanup_old_files():
    """清理超过1小时未访问的文件"""
    global running
    while running:
        current_time = time.time()
        for session_id in list(session_last_access.keys()):
            last_access = session_last_access[session_id]
            if current_time - last_access > 3600:  # 超过1小时
                session_dir = os.path.join(TEMP_DIR, session_id)
                if os.path.exists(session_dir):
                    shutil.rmtree(session_dir)
                    print(f"Deleted old session directory: {session_dir}")
                del session_last_access[session_id]
        time.sleep(60)  # 每分钟检查一次


def signal_handler(sig, frame):
    """处理信号的函数"""
    print("Stopping the application...")
    os._exit(1)  # 强制终止程序


@app.route('/', methods=['GET', 'POST'])
def index():
    try:
        if request.method == 'POST':
            # 为每个用户会话生成一个唯一的目录
            session_id = str(uuid.uuid4())
            session['session_id'] = session_id
            session_dir = os.path.join(TEMP_DIR, session_id)
            os.makedirs(session_dir, exist_ok=True)

            # 更新最后访问时间
            session_last_access[session_id] = time.time()

            # 获取上传的文件
            file1 = request.files['file1']
            file2 = request.files['file2']

            # 获取用户输入的参数
            color = request.form['color']
            file1_column = request.form['file1_column']
            file2_column = request.form['file2_column']
            file1_header_row = int(request.form['file1_header_row'])
            file2_header_row = int(request.form['file2_header_row'])

            # 保存上传的文件到会话目录
            file1_path = os.path.join(session_dir, 'file1.xlsx')
            file2_path = os.path.join(session_dir, 'file2.xlsx')
            file1.save(file1_path)
            file2.save(file2_path)

            # 获取必要的配置信息
            server_name = request.host
            application_root = '/'
            preferred_url_scheme = 'http' if not request.is_secure else 'https'

            # 异步处理文件，传递必要的配置信息
            socketio.start_background_task(process_files, file1_path, file2_path, color, file1_column, file2_column,
                                           file1_header_row, file2_header_row, session_id, server_name,
                                           application_root, preferred_url_scheme)

            return render_template('processing.html', session_id=session_id)  # 显示进度页面

        return render_template('index.html')
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"An error occurred: {str(e)}", 500


@app.route('/download/<session_id>')
def download(session_id):
    session_dir = os.path.join(TEMP_DIR, session_id)
    if not os.path.exists(session_dir):
        return "Session not found", 404

    # 更新最后访问时间
    session_last_access[session_id] = time.time()

    return render_template('download.html', session_id=session_id)


@app.route('/download/<session_id>/<filename>')
def download_file(session_id, filename):
    session_dir = os.path.join(TEMP_DIR, session_id)
    file_path = os.path.join(session_dir, filename)
    if not os.path.exists(file_path):
        return "File not found", 404

    # 更新最后访问时间
    session_last_access[session_id] = time.time()

    return send_file(file_path, as_attachment=True)


@socketio.on('join')
def on_join(data):
    session_id = data['session_id']
    join_room(session_id)


def process_files(file1, file2, color, file1_column, file2_column, file1_header_row, file2_header_row, session_id,
                  server_name, application_root, preferred_url_scheme):
    with app.app_context():
        try:
            # 设置必要的配置
            app.config['SERVER_NAME'] = server_name
            app.config['APPLICATION_ROOT'] = application_root
            app.config['PREFERRED_URL_SCHEME'] = preferred_url_scheme
            base_url = f"{preferred_url_scheme}://{server_name}{application_root}"
            # 手动创建请求上下文
            with app.test_request_context(
                    base_url=base_url
            ):
                df1 = pd.read_excel(file1, sheet_name=None, header=file1_header_row - 1)
                df2 = pd.read_excel(file2, sheet_name=None, header=file2_header_row - 1)

                wb1 = load_workbook(file1)
                wb2 = load_workbook(file2)

                color = color.lstrip('#')
                highlight_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                total_sheets = len(wb1.sheetnames)
                processed_sheets = 0

                for sheet_name1 in wb1.sheetnames:
                    ws1 = wb1[sheet_name1]
                    df1_sheet = df1[sheet_name1]

                    if file1_column not in df1_sheet.columns:
                        continue

                    column_index_file1 = df1_sheet.columns.get_loc(file1_column) + 1

                    for row1 in ws1.iter_rows(min_row=file1_header_row + 1, min_col=column_index_file1,
                                              max_col=column_index_file1):
                        for cell1 in row1:
                            cell_value1 = str(cell1.value) if cell1.value is not None else ""
                            found_match = False

                            for sheet_name2 in wb2.sheetnames:
                                ws2 = wb2[sheet_name2]
                                df2_sheet = df2[sheet_name2]

                                if file2_column not in df2_sheet.columns:
                                    continue

                                column_index_file2 = df2_sheet.columns.get_loc(file2_column) + 1
                                values_file2 = df2_sheet[file2_column].dropna().astype(str)

                                if cell_value1 in values_file2.values:
                                    matching_row = values_file2[values_file2 == cell_value1].index[0] + file2_header_row + 1

                                    for col in ws1.iter_cols(min_row=cell1.row, max_row=cell1.row, min_col=1,
                                                             max_col=ws1.max_column):
                                        for cell_to_fill in col:
                                            cell_to_fill.fill = highlight_fill

                                    for col in ws2.iter_cols(min_row=matching_row, max_row=matching_row, min_col=1,
                                                             max_col=ws2.max_column):
                                        for cell_to_fill in col:
                                            cell_to_fill.fill = highlight_fill

                                    found_match = True
                                    break

                    processed_sheets += 1
                    progress = (processed_sheets / total_sheets) * 100
                    socketio.emit('progress', {'progress': progress}, room=session_id)
                    socketio.sleep(0.1)  # 确保消息发送的正确性

                output_file1 = os.path.join(os.path.dirname(file1), 'file1_highlighted.xlsx')
                output_file2 = os.path.join(os.path.dirname(file2), 'file2_highlighted.xlsx')
                wb1.save(output_file1)
                wb2.save(output_file2)

                # 直接跳转到下载页面
                download_url = url_for('download', session_id=session_id, _external=True)
                socketio.emit('complete', {'redirect': download_url}, room=session_id)

        except Exception as e:
            import traceback
            traceback.print_exc()
            socketio.emit('error', {'message': str(e)}, room=session_id)


if __name__ == '__main__':
    # 设置信号处理函数
    signal.signal(signal.SIGINT, signal_handler)

    # 设置定时清理任务
    cleanup_thread = Thread(target=cleanup_old_files)
    cleanup_thread.start()

    try:
        app.run(debug=True)
    except KeyboardInterrupt:
        pass
    finally:
        print("Stopping the application...")
        os._exit(1)  # 强制终止程序
